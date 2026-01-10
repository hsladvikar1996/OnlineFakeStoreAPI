import json
import csv
from openpyxl import load_workbook

#Json Data Provider
def read_json_data(filename):
    with open(filename, 'r') as f:
        data = json.load(f)
    return [(item,) for item in data]

def read_excel_data(filepath, sheet_name):
    wb = load_workbook(filepath)
    sheet = wb[sheet_name]

    # Get headers from first row
    headers = [cell.value for cell in sheet[1]]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):  # skip completely empty rows
            data.append(dict(zip(headers, row)))

    return data
